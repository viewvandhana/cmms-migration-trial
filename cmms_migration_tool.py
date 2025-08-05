
import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

st.title("📥 CMMS Data Migration Tool (with Template Download)")

st.write("Upload your CMMS field rules Excel file and a data file to auto-map fields, validate types, clean data, and download templates.")

# Step 1: Load rules from Excel
def load_field_rules_from_excel(file):
    df = pd.read_excel(file)

    cmms_fields = df["Field Name"].tolist()
    field_rules = {}

    for _, row in df.iterrows():
        raw_ref = str(row.get("Reference Values", "")).strip()
        ref_values = []
        if raw_ref and any(c.isalnum() for c in raw_ref):  # only consider if there's meaningful content
         ref_values = [val.strip() for val in raw_ref.split(";") if val.strip()]
        field_rules[row["Field Name"]] = {
            "type": row["Type"],
            "required": bool(row["Required"]),
            "ref_values": ref_values
        }

    synonym_map = {}
    for _, row in df.iterrows():
        synonyms = [s.strip().lower() for s in str(row["Synonyms"]).split(";") if s.strip()]
        synonym_map[row["Field Name"]] = synonyms

    return cmms_fields, field_rules, synonym_map


# Step 2: Synonym-based mapper
def map_using_synonyms(user_columns, synonym_map):
    field_map = {}
    for col in user_columns:
        mapped = "Unmapped"
        col_lower = col.strip().lower()
        for cmms_field, synonyms in synonym_map.items():
            if col_lower == cmms_field.lower() or col_lower in synonyms:
                mapped = cmms_field
                break
        field_map[col] = mapped
    return field_map

# Step 3: Validation and cleaning
def validate_and_clean(df, field_map, field_rules):
    validation_report = []
    cleaned_df = pd.DataFrame()
    error_log = []

    for source_col, target_col in field_map.items():
        if target_col == "Unmapped" or target_col not in field_rules:
            continue

        rule = field_rules[target_col]
        field_type = rule["type"]
        required = rule["required"]
        ref_values = rule.get("ref_values", [])
        original_series = df[source_col]
        cleaned_series = original_series.copy()

        for idx, val in original_series.items():
            row_num = idx + 2  # Excel-style (assuming headers in row 1)

            # Required field check
            if required and (pd.isna(val) or val == ""):
                error_log.append({
                    "Row": row_num,
                    "Column": source_col,
                    "Issue": "Missing required value"
                })
                continue

            # Type validation
            if field_type == "Date":
                try:
                    cleaned_series[idx] = pd.to_datetime(val)
                except:
                    error_log.append({
                        "Row": row_num,
                        "Column": source_col,
                        "Issue": "Invalid date format"
                    })
                    cleaned_series[idx] = pd.NaT

            elif field_type == "Number":
                try:
                    cleaned_series[idx] = pd.to_numeric(val)
                except:
                    error_log.append({
                        "Row": row_num,
                        "Column": source_col,
                        "Issue": "Invalid number"
                    })
                    cleaned_series[idx] = pd.NA

            elif field_type == "Text":
                try:
                    cleaned_series[idx] = str(val)
                except:
                    error_log.append({
                        "Row": row_num,
                        "Column": source_col,
                        "Issue": "Invalid text"
                    })

            # Reference validation
            if ref_values and str(val).strip() not in ref_values:
                error_log.append({
                    "Row": row_num,
                    "Column": source_col,
                    "Issue": f"Value '{val}' not in reference list"
                })

        cleaned_df[target_col] = cleaned_series

    return cleaned_df, validation_report, pd.DataFrame(error_log)

# Step 4: Generate Excel template from field rules
def generate_excel_template(cmms_fields, cmms_field_rules):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CMMS Template"

    # Add headers + type hints
    for col_index, field in enumerate(cmms_fields, start=1):
        ws.cell(row=1, column=col_index, value=field)
        hint = f"({cmms_field_rules[field]['type']}){'*' if cmms_field_rules[field]['required'] else ''}"
        ws.cell(row=2, column=col_index, value=hint)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_index)].width = max(15, len(field) + 5)

    # Create hidden reference sheet
    ref_ws = wb.create_sheet("ReferenceData")
    ref_ws.sheet_state = 'hidden'
    ref_field_map = {}
    ref_col_index = 1

    for field, rule in cmms_field_rules.items():
        ref_values = rule.get("ref_values", [])
        if ref_values:
            for i, val in enumerate(ref_values):
                ref_ws.cell(row=i+1, column=ref_col_index, value=val)
            col_letter = openpyxl.utils.get_column_letter(ref_col_index)
            ref_range = f"ReferenceData!${col_letter}$1:${col_letter}${len(ref_values)}"
            ref_field_map[field] = ref_range
            ref_col_index += 1

    # Apply validations
    for col_index, field in enumerate(cmms_fields, start=1):
        rule = cmms_field_rules.get(field, {})
        col_letter = openpyxl.utils.get_column_letter(col_index)
        target_range = f"{col_letter}3:{col_letter}10"

        if field in ref_field_map:
            dv = DataValidation(type="list", formula1=f"={ref_field_map[field]}", showDropDown=True)
            dv.error = "Invalid selection"
            dv.prompt = "Choose from list"
            dv.promptTitle = f"{field}"
            dv.add(target_range)
            ws.add_data_validation(dv)

        elif rule.get("type") == "Number":
            dv = DataValidation(type="decimal", allow_blank=not rule["required"])
            dv.error = "Please enter a valid number"
            dv.prompt = "Enter a number"
            dv.promptTitle = f"{field} (Number)"
            dv.add(target_range)
            ws.add_data_validation(dv)

        elif rule.get("type") == "Date":
            dv = DataValidation(type="date", allow_blank=not rule["required"])
            dv.error = "Please enter a valid date"
            dv.prompt = "Enter a date (YYYY-MM-DD)"
            dv.promptTitle = f"{field} (Date)"
            dv.add(target_range)
            ws.add_data_validation(dv)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# Upload rules file
rules_file = st.file_uploader("Upload CMMS Field Rules Excel File", type=["xlsx"])

if rules_file:
    cmms_fields, cmms_field_rules, synonym_map = load_field_rules_from_excel(rules_file)
    st.success("✅ Field rules loaded successfully!")

    # 📥 Download Excel template button
    excel_bytes = generate_excel_template(cmms_fields, cmms_field_rules)
    st.download_button(
        label="📄 Download Excel Template",
        data=excel_bytes,
        file_name="cmms_data_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Upload data file
    uploaded_file = st.file_uploader("Upload your data file (Excel or CSV)", type=["csv", "xlsx"])

    if uploaded_file:
        if uploaded_file.name.endswith(".csv"):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)

        st.subheader("📄 Uploaded Data Preview")
        st.dataframe(df.head())

        st.subheader("🔄 Field Mapping Using Synonyms")
        user_columns = df.columns.tolist()
        mapped_fields = map_using_synonyms(user_columns, synonym_map)

        mapping_df = pd.DataFrame({
            "Your Column": list(mapped_fields.keys()),
            "Mapped To": list(mapped_fields.values())
        })
        st.dataframe(mapping_df)

        missing_required_fields = [
            field for field, rules in cmms_field_rules.items()
            if rules["required"] and field not in mapped_fields.values()
        ]
        if missing_required_fields:
            st.subheader("⚠️ Missing Required Fields in Uploaded Sheet")
            for field in missing_required_fields:
                st.error(f"Required field not found in upload: **{field}**")

        st.subheader("✅ Validation & Cleaning")
        cleaned_data, report, error_df = validate_and_clean(df, mapped_fields, cmms_field_rules)

st.subheader("🧹 Cleaned Data Preview")
st.dataframe(cleaned_data.head())

csv = cleaned_data.to_csv(index=False).encode('utf-8')
st.download_button("📥 Download Cleaned Data", csv, "cleaned_cmms_data.csv", "text/csv")

if not error_df.empty:
    st.subheader("❌ Cell-level Error Log")
    st.dataframe(error_df)

    error_csv = error_df.to_csv(index=False).encode('utf-8')
    st.download_button("📥 Download Error Log", error_csv, "cmms_error_log.csv", "text/csv")
else:
    st.success("🎉 No cell-level validation errors!")


for line in report:
    st.write("• " + line)

    st.subheader("🧹 Cleaned Data Preview")
    st.dataframe(cleaned_data.head())

    csv = cleaned_data.to_csv(index=False).encode('utf-8')
    st.download_button("📥 Download Cleaned Data", csv, "cleaned_cmms_data.csv", "text/csv")
